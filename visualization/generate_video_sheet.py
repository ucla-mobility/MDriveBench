#!/usr/bin/env python3
"""Generate an Excel index of Drive videos with infraction metadata."""
from __future__ import annotations

import csv
import json
import os
import re
import shutil
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Dict, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:  # pragma: no cover - runtime dependency bootstrap
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--user", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.worksheet.table import Table, TableStyleInfo

REMOTE = "gdrive:/Mobility Lab - Multi Agent Reasoning/Simulation Results/Interdrive - Pre Fine Tune /Videos"
OUTPUT = Path("interdrive_video_links.xlsx")
INFRACTIONS_CSV = Path("results/report/tables/infractions.csv")

RCLONE = shutil.which("rclone") or os.path.expanduser("~/.local/bin/rclone")
if not RCLONE or not Path(RCLONE).exists():
    sys.exit("rclone binary not found on PATH")

R_PREFIX = re.compile(r"^r(\d+)")


def run_rclone(*args: str) -> str:
    """Execute rclone command and return stdout."""
    result = subprocess.run(  # noqa: S603
        [RCLONE, *args],
        check=True,
        capture_output=True,
        text=True,
    )
    return result.stdout


@dataclass
class VideoItem:
    path: str
    name: str
    size: int
    file_id: str | None

    @property
    def basename(self) -> str:
        return PurePosixPath(self.path).name

    @property
    def route_key(self) -> str:
        stem = PurePosixPath(self.basename).stem
        if not stem.startswith("r"):
            return stem
        marker = stem.find("_10_")
        return stem[:marker] if marker != -1 else stem


@dataclass
class VideoRow(VideoItem):
    hyperlink: str
    infraction_types: List[str]
    infraction_details: List[str]
    infraction_total: int
    per_mode: Dict[str, Dict[str, int]]

    @property
    def has_infractions(self) -> bool:
        return bool(self.infraction_types)


def fetch_video_items() -> List[VideoItem]:
    payload = run_rclone(
        "lsjson",
        "--files-only",
        "-R",
        REMOTE,
    )
    try:
        objects = json.loads(payload)
    except json.JSONDecodeError as exc:  # pragma: no cover - defensive
        raise RuntimeError("Failed to parse rclone lsjson output") from exc

    items: List[VideoItem] = []
    for obj in objects:
        if obj.get("IsDir"):
            continue
        rel_path = obj.get("Path") or obj.get("Name")
        if not rel_path:
            continue
        items.append(
            VideoItem(
                path=rel_path,
                name=obj.get("Name", rel_path),
                size=int(obj.get("Size", 0)),
                file_id=obj.get("ID"),
            )
        )
    return items


def ensure_hyperlink(item: VideoItem) -> str:
    if item.file_id:
        return f"https://drive.google.com/file/d/{item.file_id}/view?usp=drive_link"
    remote_path = f"{REMOTE}/{item.path}".replace("//", "/")
    return run_rclone("link", remote_path).strip()


def natural_sort_key(item: VideoItem):
    match = R_PREFIX.match(item.basename)
    if match:
        number = int(match.group(1))
        return (0, number, item.basename.lower(), item.path.lower())
    return (1, item.basename.lower(), item.path.lower())


def load_infractions() -> Dict[str, Dict[str, Dict[str, int]]]:
    if not INFRACTIONS_CSV.exists():
        return {}
    data: Dict[str, Dict[str, Dict[str, int]]] = defaultdict(lambda: defaultdict(dict))
    with INFRACTIONS_CSV.open(newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            route = row["Route"].strip()
            mode = row["Mode"].strip()
            infraction = row["Infraction"].strip()
            count_raw = row["Count"].strip()
            try:
                count = int(float(count_raw))
            except ValueError:
                count = 0
            mode_bucket = data[route][mode]
            mode_bucket[infraction] = mode_bucket.get(infraction, 0) + count
    return data


def enrich(items: List[VideoItem], infractions: Dict[str, Dict[str, Dict[str, int]]]) -> List[VideoRow]:
    enriched: List[VideoRow] = []
    for item in sorted(items, key=natural_sort_key):
        route_key = item.route_key
        per_mode = infractions.get(route_key, {})
        type_counts: Dict[str, int] = defaultdict(int)
        detail_lines: List[str] = []
        for mode in sorted(per_mode):
            events = per_mode[mode]
            if not events:
                continue
            sorted_events = sorted(events.items())
            detail = ", ".join(f"{inf_type} ({count})" for inf_type, count in sorted_events)
            detail_lines.append(f"{mode}: {detail}")
            for inf_type, count in sorted_events:
                type_counts[inf_type] += count
        record = VideoRow(
            path=item.path,
            name=item.name,
            size=item.size,
            file_id=item.file_id,
            hyperlink=ensure_hyperlink(item),
            infraction_types=sorted(type_counts),
            infraction_details=detail_lines,
            infraction_total=sum(type_counts.values()),
            per_mode=per_mode,
        )
        enriched.append(record)
    return enriched


def write_videos_sheet(wb: Workbook, rows: List[VideoRow]) -> None:
    ws = wb.active
    ws.title = "Videos"
    headers = [
        "Video",
        "Size (bytes)",
        "Has Infractions",
        "Infraction Count",
        "Infraction Types",
        "Infraction Details",
    ]
    ws.append(headers)
    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for cell in ws[1]:
        cell.font = header_font

    for idx, row in enumerate(rows, start=2):
        hyperlink = f'=HYPERLINK("{row.hyperlink}","{row.basename}")'
        ws.cell(row=idx, column=1).value = hyperlink
        ws.cell(row=idx, column=2, value=row.size)
        ws.cell(row=idx, column=3, value="Yes" if row.has_infractions else "No")
        ws.cell(row=idx, column=4, value=row.infraction_total)
        types_text = "\n".join(row.infraction_types)
        details_text = "\n".join(row.infraction_details)
        type_cell = ws.cell(row=idx, column=5, value=types_text)
        detail_cell = ws.cell(row=idx, column=6, value=details_text)
        type_cell.alignment = wrap_alignment
        detail_cell.alignment = wrap_alignment

    ws.freeze_panes = "A2"
    last_row = len(rows) + 1
    table = Table(displayName="VideoIndex", ref=f"A1:F{last_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 60


def write_infractions_sheet(wb: Workbook, rows: List[VideoRow]) -> None:
    flat_rows = []
    for row in rows:
        if not row.per_mode:
            continue
        for mode, events in sorted(row.per_mode.items()):
            for inf_type, count in sorted(events.items()):
                flat_rows.append((row.basename, row.route_key, mode, inf_type, count))

    if not flat_rows:
        return

    ws = wb.create_sheet("Infractions")
    headers = ["Video", "Route", "Mode", "Infraction", "Count"]
    ws.append(headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for row in flat_rows:
        ws.append(row)

    ws.freeze_panes = "A2"
    last_row = len(flat_rows) + 1
    table = Table(displayName="VideoInfractions", ref=f"A1:E{last_row}")
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12


def main() -> None:
    items = fetch_video_items()
    infractions = load_infractions()
    rows = enrich(items, infractions)

    wb = Workbook()
    write_videos_sheet(wb, rows)
    write_infractions_sheet(wb, rows)
    wb.save(OUTPUT)
    print(f"Wrote {len(rows)} videos to {OUTPUT}")


if __name__ == "__main__":
    main()
